from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from . import _text


def write(d: dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = d["doc_type"]

    bold = Font(bold=True, size=14)
    ws["A1"] = d["title"]
    ws["A1"].font = bold
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for line in _text.render_lines(d)[1:]:
        ws.cell(row=row, column=1, value=line if line else None)
        if line:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

    if "items" in d and d["items"]:
        row += 1
        rows = _text.items_table(d["items"])
        for r_offset, src_row in enumerate(rows):
            for c, val in enumerate(src_row, 1):
                cell = ws.cell(row=row + r_offset, column=c, value=val)
                if r_offset == 0:
                    cell.font = Font(bold=True)

    for col_idx, width in enumerate([6, 36, 6, 8, 12, 14, 12], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(str(path))
